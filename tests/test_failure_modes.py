"""Failure-mode regression tests — synthetic PDFs only, no client data needed.

Run: python tests/test_failure_modes.py   (any machine, no samples/ required,
NO real API key required — every Gemini call here is faked locally).

Covers every way a statement PDF could vanish WITHOUT leaving a row in the
Excel, plus the Version-2 API failure modes:
  1. a password-protected PDF            -> must write a visible FAILED row
  2. a scanned PDF with NO API key set   -> the flag must say "open api_key.txt"
  3. a scanned PDF with a REJECTED key   -> the flag must name the key problem
  4. the internet being down             -> the flag must name the network
  5. the API being busy (HTTP 429)       -> must retry, then succeed
  6. a scanned PDF read successfully     -> full end-to-end: image page ->
     (mock) Gemini JSON -> UBS parser -> correct numbers + snapshot written
"""
import io
import json
import os
import sys
import tempfile
import urllib.error
import urllib.request
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import fitz
from openpyxl import load_workbook

import config
from shared.readers import gemini_ocr

COL_SOURCE, COL_FLAGS = 9, 12

_REAL_URLOPEN = urllib.request.urlopen


def _tmp_env() -> Path:
    """Point all outputs at a throwaway folder so tests never touch real files."""
    tmp = Path(tempfile.mkdtemp(prefix="pinnacle_test_"))
    config.MASTER_WORKBOOK = tmp / "nav_master.xlsx"
    config.SNAPSHOT_DIR = tmp / "snapshots"
    config.verification_docx = lambda bank: tmp / f"{bank}_verification.docx"
    config.API_KEY_FILE = tmp / "api_key.txt"          # never the real key file
    os.environ.pop("GEMINI_API_KEY", None)
    urllib.request.urlopen = _REAL_URLOPEN             # undo any leftover mock
    gemini_ocr._RETRY_WAITS = (0, 0)                   # don't sleep in tests
    return tmp


def _make_pdf(path: Path, text: str = "", encrypted: bool = False) -> None:
    doc = fitz.open()
    page = doc.new_page()
    if text:
        page.insert_text((72, 72), text)
    kwargs = {}
    if encrypted:
        kwargs = dict(encryption=fitz.PDF_ENCRYPT_AES_256,
                      user_pw="secret", owner_pw="secret")
    doc.save(str(path), **kwargs)
    doc.close()


def _make_scanned_pdf(path: Path, rows) -> None:
    """A PDF whose single page is ONE IMAGE (no text layer) showing `rows`."""
    src = fitz.open()
    page = src.new_page()
    for i, text in enumerate(rows):
        page.insert_text((72, 90 + i * 24), text, fontsize=10)
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    src.close()
    doc = fitz.open()
    page = doc.new_page()
    page.insert_image(page.rect, stream=pix.tobytes("png"))
    doc.save(str(path))
    doc.close()


class _FakeResp:
    def __init__(self, payload: dict):
        self._data = json.dumps(payload).encode("utf-8")

    def read(self):
        return self._data

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False


def _gemini_payload(lines) -> dict:
    """What the real API returns, wrapping our fake transcription."""
    return {"candidates": [{"content": {"parts": [
        {"text": json.dumps({"lines": lines})}]}}]}


def _flags_of(bank: str, source_name: str):
    wb = load_workbook(config.MASTER_WORKBOOK)
    ws = wb[bank]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=COL_SOURCE).value or "") == source_name:
            return str(ws.cell(row=r, column=COL_FLAGS).value or "")
    return None


UBS_ROWS = [
    "Total assets",
    "Portfolio number 546-123456-02",
    "Valued in USD",
    "Assets as of 28.02.2026",
    "Asset class Market value Accrued interest Total",
    "Liquidity 2 866 000 1 500 2 867 500",
    "Bonds 5 000 000 0 5 000 000",
    "Gross assets 14 500 000 1 500 14 501 500",
    "Liabilities -2 450 000 0 -2 450 000",
    "Net assets 12 050 000 1 500 12 051 500",
]
# What the (mock) API sends back: one entry per visual row, each column value
# its own cell. The first entry uses the plain-"text" form on purpose, to
# cover the fallback for a model that ignores the cells instruction.
UBS_LINES = [
    {"y": 8, "text": "Total assets"},
    {"y": 12, "cells": ["Portfolio number 546-123456-02"]},
    {"y": 16, "cells": ["Valued in USD"]},
    {"y": 20, "cells": ["Assets as of 28.02.2026"]},
    {"y": 24, "cells": ["Asset class", "Market value", "Accrued interest", "Total"]},
    {"y": 28, "cells": ["Liquidity", "2 866 000", "1 500", "2 867 500"]},
    {"y": 32, "cells": ["Bonds", "5 000 000", "0", "5 000 000"]},
    {"y": 36, "cells": ["Gross assets", "14 500 000", "1 500", "14 501 500"]},
    {"y": 40, "cells": ["Liabilities", "-2 450 000", "0", "-2 450 000"]},
    {"y": 44, "cells": ["Net assets", "12 050 000", "1 500", "12 051 500"]},
]


def test_encrypted_pdf_writes_failed_row(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "locked.pdf"
    _make_pdf(pdf, text="hello", encrypted=True)
    results = process_pdf("UBS", pdf)            # must NOT raise
    flags = _flags_of("UBS", "locked.pdf")
    assert flags is not None, "no row written for a password-protected PDF"
    assert "password" in flags.lower(), f"flag doesn't name the cause: {flags}"
    assert all(r.failed for r in results), \
        "failed result not marked failed — the runner would skip retrying it"
    print("PASS  encrypted PDF -> visible FAILED row naming the password")


def test_scan_without_key_names_api_key_file(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "scan_no_key.pdf"
    _make_scanned_pdf(pdf, UBS_ROWS)             # no GEMINI_API_KEY, no file
    process_pdf("UBS", pdf)                      # must NOT raise
    flags = _flags_of("UBS", "scan_no_key.pdf") or ""
    assert "api_key.txt" in flags, \
        f"flag doesn't tell her to paste the key into api_key.txt: {flags}"
    print("PASS  scan without an API key -> flag says open api_key.txt")


def test_rejected_key_names_the_cause(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "scan_bad_key.pdf"
    _make_scanned_pdf(pdf, UBS_ROWS)
    os.environ["GEMINI_API_KEY"] = "bad-key-for-test"

    def _reject(req, timeout=None):
        raise urllib.error.HTTPError(req.full_url, 403, "Forbidden", None,
                                     io.BytesIO(b'{"error":"key invalid"}'))

    urllib.request.urlopen = _reject
    try:
        process_pdf("UBS", pdf)                  # must NOT raise
    finally:
        urllib.request.urlopen = _REAL_URLOPEN
        os.environ.pop("GEMINI_API_KEY", None)
    flags = _flags_of("UBS", "scan_bad_key.pdf") or ""
    assert "key" in flags.lower() and "403" in flags, \
        f"flag doesn't name the rejected key: {flags}"
    print("PASS  rejected API key -> flag names the key problem")


def test_network_down_names_the_cause(tmp: Path) -> None:
    from shared.process import process_pdf
    pdf = tmp / "scan_offline.pdf"
    _make_scanned_pdf(pdf, UBS_ROWS)
    os.environ["GEMINI_API_KEY"] = "any-key"

    def _offline(req, timeout=None):
        raise urllib.error.URLError(OSError("no route to host"))

    urllib.request.urlopen = _offline
    try:
        process_pdf("UBS", pdf)                  # must NOT raise
    finally:
        urllib.request.urlopen = _REAL_URLOPEN
        os.environ.pop("GEMINI_API_KEY", None)
    flags = _flags_of("UBS", "scan_offline.pdf") or ""
    assert "internet" in flags.lower() or "reach" in flags.lower(), \
        f"flag doesn't name the network problem: {flags}"
    print("PASS  internet down -> flag says check the connection")


def test_api_busy_retries_then_succeeds(tmp: Path) -> None:
    from banks.UBS.parser import parse
    pdf = tmp / "scan_busy.pdf"
    _make_scanned_pdf(pdf, UBS_ROWS)
    os.environ["GEMINI_API_KEY"] = "any-key"
    calls = {"n": 0}

    def _busy_then_ok(req, timeout=None):
        calls["n"] += 1
        if calls["n"] == 1:
            raise urllib.error.HTTPError(req.full_url, 429, "Too Many", None,
                                         io.BytesIO(b'{"error":"rate limit"}'))
        return _FakeResp(_gemini_payload(UBS_LINES))

    urllib.request.urlopen = _busy_then_ok
    try:
        res = parse(pdf)[0]
    finally:
        urllib.request.urlopen = _REAL_URLOPEN
        os.environ.pop("GEMINI_API_KEY", None)
    assert calls["n"] == 2, f"expected 1 retry, saw {calls['n']} calls"
    assert res.net_nav == 12_050_000, res.net_nav
    print("PASS  API busy (429) -> retried automatically and succeeded")


def test_scanned_statement_end_to_end(tmp: Path) -> None:
    """The core Version-2 proof: an image-only UBS statement goes through the
    (mock) Gemini API and comes out with the exact figures, a row in the
    Excel, and a snapshot PNG for the verification doc."""
    from shared.process import process_pdf
    pdf = tmp / "ubs_scanned.pdf"
    _make_scanned_pdf(pdf, UBS_ROWS)
    os.environ["GEMINI_API_KEY"] = "any-key"

    urllib.request.urlopen = (
        lambda req, timeout=None: _FakeResp(_gemini_payload(UBS_LINES)))
    try:
        results = process_pdf("UBS", pdf)        # full pipeline
    finally:
        urllib.request.urlopen = _REAL_URLOPEN
        os.environ.pop("GEMINI_API_KEY", None)

    res = results[0]
    assert not res.failed, f"pipeline failed: {res.flags}"
    assert res.account_no == "546-123456-02", res.account_no
    assert res.currency == "USD", res.currency
    assert res.gross_nav == 14_500_000, res.gross_nav
    assert res.net_nav == 12_050_000, res.net_nav
    assert res.liquidity == 2_866_000, res.liquidity
    assert res.liabilities == -2_450_000, res.liabilities
    assert _flags_of("UBS", "ubs_scanned.pdf") is not None, "no Excel row"
    snaps = list((config.SNAPSHOT_DIR / "UBS").glob("ubs_scanned__*.png"))
    assert snaps, "no snapshot PNG rendered for the verification doc"
    print("PASS  scanned UBS statement -> API OCR -> exact figures + snapshot")


def test_ubs_single_portfolio_statement(tmp: Path) -> None:
    """UBS also exports ONE PDF per portfolio: same asset-class table on the
    overview page but with NO 'Portfolio NN' heading (colleague's real files,
    2026-07-07). The parser must read that table instead of coming back empty."""
    from banks.UBS.parser import parse
    pdf = tmp / "ubs_single.pdf"
    doc = fitz.open()
    page = doc.new_page()
    for i, text in enumerate(UBS_ROWS):
        page.insert_text((72, 90 + i * 24), text, fontsize=10)
    doc.save(str(pdf))
    doc.close()

    res = parse(pdf)[0]
    assert res.account_no == "546-123456-02", res.account_no
    assert res.currency == "USD", res.currency
    assert res.gross_nav == 14_500_000, res.gross_nav
    assert res.net_nav == 12_050_000, res.net_nav
    assert res.liquidity == 2_866_000, res.liquidity
    assert res.liabilities == -2_450_000, res.liabilities
    print("PASS  UBS one-portfolio-per-PDF statement -> table read directly")


def test_ubs_two_tables_takes_first(tmp: Path) -> None:
    """When a heading-less UBS page shows SEVERAL asset tables, the client's own
    portfolio is printed first (per the suffix) — read the FIRST table only and
    never mix rows across tables."""
    from banks.UBS.parser import parse
    pdf = tmp / "ubs_two_tables.pdf"
    doc = fitz.open()
    page = doc.new_page()
    rows = [
        "Total assets",
        "Portfolio number 546-123456-03",
        "Valued in USD",
        "Liquidity 1 000 000 0 1 000 000",
        "Equities 9 000 000 0 9 000 000",
        "Gross assets 10 000 000 0 10 000 000",
        "Liabilities -1 000 000 0 -1 000 000",
        "Net assets 9 000 000 0 9 000 000",
        # a second portfolio's table further down, no heading either
        "Liquidity 77 0 77",
        "Gross assets 99 0 99",
        "Net assets 98 0 98",
    ]
    for i, text in enumerate(rows):
        page.insert_text((72, 90 + i * 24), text, fontsize=10)
    doc.save(str(pdf))
    doc.close()

    res = parse(pdf)[0]
    assert res.gross_nav == 10_000_000, res.gross_nav
    assert res.net_nav == 9_000_000, res.net_nav
    assert res.liquidity == 1_000_000, res.liquidity
    assert res.liabilities == -1_000_000, res.liabilities
    assert any("FIRST" in f for f in res.flags), res.flags
    print("PASS  UBS several heading-less tables -> first table only, flagged")


def main() -> None:
    failures = 0
    for test in (test_encrypted_pdf_writes_failed_row,
                 test_scan_without_key_names_api_key_file,
                 test_rejected_key_names_the_cause,
                 test_network_down_names_the_cause,
                 test_api_busy_retries_then_succeeds,
                 test_scanned_statement_end_to_end,
                 test_ubs_single_portfolio_statement,
                 test_ubs_two_tables_takes_first):
        try:
            test(_tmp_env())
        except Exception as exc:
            failures += 1
            print(f"FAIL  {test.__name__}: {type(exc).__name__}: {exc}")
    urllib.request.urlopen = _REAL_URLOPEN
    if failures:
        sys.exit(f"{failures} failure-mode test(s) failed")
    print("All failure-mode tests passed.")


if __name__ == "__main__":
    main()
